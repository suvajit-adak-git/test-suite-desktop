import openpyxl
from typing import List, Dict
import re
from fastapi import HTTPException


def validate_tc_traceability(file_path: str) -> Dict:
    """
    Validates TC traceability across Excel workbook.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Dictionary with summary and validation results
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to open workbook: {str(e)}")
    
    if 'General' not in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=400, detail="'General' sheet not found in workbook")
    
    general_sheet = wb['General']
    results = []
    
    # Find the header row and column indices
    req_id_col = None
    tc_col = None
    header_row = None
    
    # Search for headers in first 20 rows
    for row_idx in range(1, 21):
        for col_idx, cell in enumerate(general_sheet[row_idx], start=1):
            cell_value = str(cell.value).strip() if cell.value else ""
            if cell_value == "Requirements ID":
                req_id_col = col_idx
                header_row = row_idx
            elif cell_value == "TC":
                tc_col = col_idx
    
    if not req_id_col or not tc_col:
        wb.close()
        raise HTTPException(
            status_code=400, 
            detail="Could not find 'Requirements ID' or 'TC' columns in General sheet"
        )
    
    # Extract requirements starting from row after header
    for row_idx in range(header_row + 1, general_sheet.max_row + 1):
        req_id_cell = general_sheet.cell(row_idx, req_id_col)
        tc_cell = general_sheet.cell(row_idx, tc_col)
        
        req_id = str(req_id_cell.value).strip() if req_id_cell.value else ""
        tc_value = str(tc_cell.value).strip() if tc_cell.value else ""
        
        # Skip empty rows
        if not req_id or req_id == "None":
            continue
        
        result = {
            "requirement_id": req_id,
            "expected_tcs": [],
            "found_in_sheets": [],
            "status": "pass",
            "error": None
        }
        
        # Check if TC is N/A or empty
        if not tc_value or tc_value.upper() in ["N/A", "NONE"]:
            result["status"] = "fail"
            result["error"] = f"TC column is N/A or empty for requirement {req_id}"
            results.append(result)
            continue
        
        # Parse TC values (e.g., "TC_1, TC_2, TC_3")
        expected_tcs = [tc.strip() for tc in re.split(r'[,;]', tc_value)]
        result["expected_tcs"] = expected_tcs
        
        # Search for requirement ID in all sheets except General
        found_sheets = []
        for sheet_name in wb.sheetnames:
            if sheet_name == "General":
                continue
            
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    cell_value = str(cell.value).strip() if cell.value else ""
                    if req_id in cell_value:
                        found_sheets.append(sheet_name)
                        break
                if sheet_name in found_sheets:
                    break
        
        result["found_in_sheets"] = found_sheets
        
        # Validate: Check if requirement appears in ALL expected TC sheets
        missing_sheets = set(expected_tcs) - set(found_sheets)
        extra_sheets = set(found_sheets) - set(expected_tcs)
        
        if missing_sheets:
            result["status"] = "fail"
            result["error"] = f"Requirement {req_id} not found in expected sheets: {', '.join(missing_sheets)}"
        elif extra_sheets:
            result["status"] = "fail"
            result["error"] = f"Requirement {req_id} found in unexpected sheets: {', '.join(extra_sheets)}"
        
        results.append(result)
    
    wb.close()
    
    # Calculate summary
    total = len(results)
    passed = sum(1 for r in results if r["status"] == "pass")
    failed = total - passed
    
    return {
        "summary": {
            "total_requirements": total,
            "passed": passed,
            "failed": failed
        },
        "results": results
    }
