import shutil
from pathlib import Path
from typing import Dict, Any, List
import pandas as pd
import openpyxl
from fastapi import UploadFile, HTTPException
from app.utils.common import cell_fill_rgb

def save_upload_file(upload_file: UploadFile, dest: Path) -> None:
    with dest.open("wb") as buffer:
        shutil.copyfileobj(upload_file.file, buffer)

def extract_from_excel_file(path: Path, max_preview_rows: int = 100) -> Dict[str, Any]:
    try:
        df = pd.read_excel(path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read excel file: {e}")

    headers = list(df.columns)
    nrows = int(len(df))
    preview_df = df.head(max_preview_rows).fillna("")
    preview = preview_df.to_dict(orient="records")
    sample_values = []
    if headers:
        first_col = headers[0]
        sample_values = [str(x).strip() for x in df[first_col].dropna().astype(str).unique().tolist()]

    return {"filename": path.name, "headers": headers, "nrows": nrows, "preview": preview, "sample_values": sample_values}

def extract_from_csv_file(path: Path, max_preview_rows: int = 100) -> Dict[str, Any]:
    df = None
    try:
        df = pd.read_csv(path, engine="python", sep=None)
    except Exception:
        for sep in [",", ";", "\t", "|"]:
            try:
                df = pd.read_csv(path, sep=sep)
                break
            except Exception:
                df = None
    if df is None:
        raise HTTPException(status_code=400, detail="Failed to parse CSV file (invalid format or delimiter).")

    df = df.dropna(how="all")
    df = df[~(df.astype(str).apply(lambda x: "".join(x).strip(), axis=1) == "")]

    headers = list(df.columns)
    nrows = int(len(df))
    preview_df = df.head(max_preview_rows).fillna("")
    preview = preview_df.to_dict(orient="records")
    sample_values = []
    if headers:
        first_col = headers[0]
        sample_values = [str(x).strip() for x in df[first_col].dropna().astype(str).unique().tolist()]

    return {"filename": path.name, "headers": headers, "nrows": nrows, "preview": preview, "sample_values": sample_values}

def extract_hyperlinks_with_versions_from_path(file_path: str, sheet_name: str = "Test Scenario Remarks") -> List[Dict[str, Any]]:
    """
    Enhanced extraction that:
    1. Extracts from Test Scenario Remarks (primary sheet)
    2. Also extracts hyperlinks from Test Case Remarks
    3. Merges both, using max(version) when file appears in both
    4. Flags inter-sheet version conflicts
    """
    p = Path(file_path)
    if not p.exists() or not p.is_file():
        raise HTTPException(status_code=404, detail=f"File not found: {file_path}")

    try:
        workbook = openpyxl.load_workbook(p, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to open workbook: {e}")

    # ========== PART 1: Extract from Test Scenario Remarks (PRIMARY) ==========
    test_scenario_results = {}  # Dict[normalized_filename, dict]
    
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Section 1: "REVIEWED AGAINST"
        doc_name_col = None
        doc_revision_col = None
        reviewed_against_header_row = None

        for row_idx in range(1, min(21, sheet.max_row + 1)):
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row_idx, col_idx).value
                if cell_value:
                    cell_value_str = str(cell_value).strip().lower()
                    if cell_value_str == "document name":
                        doc_name_col = col_idx
                        reviewed_against_header_row = row_idx
                    if cell_value_str == "document revision":
                        doc_revision_col = col_idx
            if doc_name_col and doc_revision_col:
                break

        if doc_name_col and doc_revision_col and reviewed_against_header_row:
            for row_idx in range(reviewed_against_header_row + 1, min(reviewed_against_header_row + 5, sheet.max_row + 1)):
                doc_name_cell = sheet.cell(row_idx, doc_name_col)
                doc_revision_cell = sheet.cell(row_idx, doc_revision_col)
                
                doc_name = doc_name_cell.value
                if not doc_name:
                    break
                
                hyperlink_url = None
                if getattr(doc_name_cell, "hyperlink", None):
                    try:
                        hyperlink_url = doc_name_cell.hyperlink.target
                    except Exception:
                        pass
                
                doc_name_str = str(doc_name).strip()
                doc_revision_str = str(doc_revision_cell.value).strip() if doc_revision_cell.value else None
                
                # Normalize filename for matching
                from app.utils.common import normalize_filename_for_match
                norm_name = normalize_filename_for_match(doc_name_str)
                
                test_scenario_results[norm_name] = {
                    "filename": doc_name_str,
                    "hyperlink": hyperlink_url,
                    "version_closed": doc_revision_str,
                    "row": row_idx,
                    "source_sheet": "Test Scenario Remarks"
                }

        # Section 2: "ARTIFACT(S) UNDER REVIEW"
        filename_col = None
        version_closed_col = None
        header_row = None

        for row_idx in range(1, min(21, sheet.max_row + 1)):
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row_idx, col_idx).value
                if cell_value:
                    cell_value_str = str(cell_value).strip().lower()
                    if "filename" in cell_value_str:
                        filename_col = col_idx
                        header_row = row_idx
                    if "version on which" in cell_value_str and "closed" in cell_value_str:
                        version_closed_col = col_idx
            if filename_col and version_closed_col:
                break

        if filename_col and version_closed_col and header_row:
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                filename_cell = sheet.cell(row_idx, filename_col)
                version_cell = sheet.cell(row_idx, version_closed_col)

                # Check for green background (end marker)
                color = cell_fill_rgb(filename_cell)
                if color and len(color) >= 6:
                    hexstr = color[-8:] if len(color) >= 8 else color
                    try:
                        if len(hexstr) == 8:
                            r = int(hexstr[2:4], 16)
                            g = int(hexstr[4:6], 16)
                            b = int(hexstr[6:8], 16)
                        elif len(hexstr) == 6:
                            r = int(hexstr[0:2], 16)
                            g = int(hexstr[2:4], 16)
                            b = int(hexstr[4:6], 16)
                        else:
                            r = g = b = 0
                    except Exception:
                        r = g = b = 0

                    if g > r and g > b and g > 100:
                        break

                filename = filename_cell.value
                if not filename:
                    continue

                hyperlink_url = None
                if getattr(filename_cell, "hyperlink", None):
                    try:
                        hyperlink_url = filename_cell.hyperlink.target
                    except Exception:
                        pass

                filename_str = str(filename).strip()
                version_closed_str = str(version_cell.value).strip() if version_cell.value else None

                from app.utils.common import normalize_filename_for_match
                norm_name = normalize_filename_for_match(filename_str)
                
                test_scenario_results[norm_name] = {
                    "filename": filename_str,
                    "hyperlink": hyperlink_url,
                    "version_closed": version_closed_str,
                    "row": row_idx,
                    "source_sheet": "Test Scenario Remarks"
                }

    # ========== PART 2: Extract from Test Case Remarks ==========
    test_case_results = {}  # Dict[normalized_filename, dict]
    
    if "Test Case Remarks" in workbook.sheetnames:
        tc_sheet = workbook["Test Case Remarks"]
        
        # Find "ARTIFACT(S) UNDER REVIEW" section
        tc_filename_col = None
        tc_version_closed_col = None
        tc_header_row = None

        for row_idx in range(1, min(21, tc_sheet.max_row + 1)):
            for col_idx in range(1, tc_sheet.max_column + 1):
                cell_value = tc_sheet.cell(row_idx, col_idx).value
                if cell_value:
                    cell_value_str = str(cell_value).strip().lower()
                    if "filename" in cell_value_str:
                        tc_filename_col = col_idx
                        tc_header_row = row_idx
                    if "version on which" in cell_value_str and "closed" in cell_value_str:
                        tc_version_closed_col = col_idx
            if tc_filename_col and tc_version_closed_col:
                break

        # Extract from ARTIFACT(S) UNDER REVIEW section
        if tc_filename_col and tc_version_closed_col and tc_header_row:
            for row_idx in range(tc_header_row + 1, tc_sheet.max_row + 1):
                filename_cell = tc_sheet.cell(row_idx, tc_filename_col)
                version_cell = tc_sheet.cell(row_idx, tc_version_closed_col)

                # Check for green background
                color = cell_fill_rgb(filename_cell)
                if color and len(color) >= 6:
                    hexstr = color[-8:] if len(color) >= 8 else color
                    try:
                        if len(hexstr) == 8:
                            r = int(hexstr[2:4], 16)
                            g = int(hexstr[4:6], 16)
                            b = int(hexstr[6:8], 16)
                        elif len(hexstr) == 6:
                            r = int(hexstr[0:2], 16)
                            g = int(hexstr[2:4], 16)
                            b = int(hexstr[4:6], 16)
                        else:
                            r = g = b = 0
                    except Exception:
                        r = g = b = 0

                    if g > r and g > b and g > 100:
                        break

                filename = filename_cell.value
                if not filename:
                    continue

                hyperlink_url = None
                if getattr(filename_cell, "hyperlink", None):
                    try:
                        hyperlink_url = filename_cell.hyperlink.target
                    except Exception:
                        pass

                filename_str = str(filename).strip()
                version_closed_str = str(version_cell.value).strip() if version_cell.value else None

                from app.utils.common import normalize_filename_for_match
                norm_name = normalize_filename_for_match(filename_str)
                
                test_case_results[norm_name] = {
                    "filename": filename_str,
                    "hyperlink": hyperlink_url,
                    "version_closed": version_closed_str,
                    "row": row_idx,
                    "source_sheet": "Test Case Remarks"
                }
        
        # Scan "Items" section for hyperlinks (below green line in ARTIFACT(S) UNDER REVIEW)
        # Look for "Items" header
        items_col = None
        items_header_row = None
        
        for row_idx in range(1, min(30, tc_sheet.max_row + 1)):
            cell_value = tc_sheet.cell(row_idx, 1).value  # Column A
            if cell_value:
                cell_value_str = str(cell_value).strip().lower()
                if cell_value_str == "items":
                    items_col = 1
                    items_header_row = row_idx
                    break
        
        if items_header_row:
            for row_idx in range(items_header_row + 1, tc_sheet.max_row + 1):
                item_cell = tc_sheet.cell(row_idx, items_col)
                
                # Check for green background (end marker)
                color = cell_fill_rgb(item_cell)
                if color and len(color) >= 6:
                    hexstr = color[-8:] if len(color) >= 8 else color
                    try:
                        if len(hexstr) == 8:
                            r = int(hexstr[2:4], 16)
                            g = int(hexstr[4:6], 16)
                            b = int(hexstr[6:8], 16)
                        elif len(hexstr) == 6:
                            r = int(hexstr[0:2], 16)
                            g = int(hexstr[2:4], 16)
                            b = int(hexstr[4:6], 16)
                        else:
                            r = g = b = 0
                    except Exception:
                        r = g = b = 0

                    if g > r and g > b and g > 100:
                        break
                
                item_value = item_cell.value
                if not item_value:
                    continue
                
                # Check if cell has hyperlink
                hyperlink_url = None
                if getattr(item_cell, "hyperlink", None):
                    try:
                        hyperlink_url = item_cell.hyperlink.target
                    except Exception:
                        pass
                
                # Only process if there's a hyperlink
                if hyperlink_url:
                    item_str = str(item_value).strip()
                    
                    from app.utils.common import normalize_filename_for_match
                    norm_name = normalize_filename_for_match(item_str)
                    
                    # Items section doesn't have version info, set as None
                    test_case_results[norm_name] = {
                        "filename": item_str,
                        "hyperlink": hyperlink_url,
                        "version_closed": None,
                        "row": row_idx,
                        "source_sheet": "Test Case Remarks"
                    }

    # ========== PART 3: Merge Results ==========
    from app.utils.common import extract_int_from_version
    
    final_results = []
    merged_keys = set()
    
    # Process all files
    all_norm_names = set(test_scenario_results.keys()) | set(test_case_results.keys())
    
    for norm_name in all_norm_names:
        tc_entry = test_case_results.get(norm_name)
        ts_entry = test_scenario_results.get(norm_name)
        
        if tc_entry and ts_entry:
            # File exists in BOTH sheets
            tc_version_str = tc_entry.get("version_closed")
            ts_version_str = ts_entry.get("version_closed")
            
            tc_version_int = extract_int_from_version(tc_version_str)
            ts_version_int = extract_int_from_version(ts_version_str)
            
            # Use max(X, Y) logic
            if tc_version_int is not None and ts_version_int is not None:
                if tc_version_int != ts_version_int:
                    # Version mismatch between sheets
                    max_version_int = max(tc_version_int, ts_version_int)
                    max_version_str = tc_version_str if tc_version_int > ts_version_int else ts_version_str
                    
                    # Use Test Scenario Remarks as precedence for metadata
                    final_results.append({
                        "filename": ts_entry["filename"],
                        "hyperlink": ts_entry["hyperlink"],
                        "version_closed": max_version_str,
                        "row": ts_entry["row"],
                        "inter_sheet_conflict": True,
                        "conflict_comment": f"Version mismatch between sheets: Test Case Remarks has {tc_version_str}, Test Scenario Remarks has {ts_version_str}. Using max: {max_version_str}"
                    })
                else:
                    # Same version in both sheets - Test Scenario Remarks takes precedence
                    final_results.append({
                        "filename": ts_entry["filename"],
                        "hyperlink": ts_entry["hyperlink"],
                        "version_closed": ts_entry["version_closed"],
                        "row": ts_entry["row"]
                    })
            else:
                # Use Test Scenario Remarks (has precedence)
                final_results.append({
                    "filename": ts_entry["filename"],
                    "hyperlink": ts_entry["hyperlink"],
                    "version_closed": ts_entry["version_closed"],
                    "row": ts_entry["row"]
                })
        
        elif ts_entry:
            # Only in Test Scenario Remarks
            final_results.append({
                "filename": ts_entry["filename"],
                "hyperlink": ts_entry["hyperlink"],
                "version_closed": ts_entry["version_closed"],
                "row": ts_entry["row"]
            })
        
        elif tc_entry:
            # Only in Test Case Remarks - add to list
            final_results.append({
                "filename": tc_entry["filename"],
                "hyperlink": tc_entry["hyperlink"],
                "version_closed": tc_entry["version_closed"],
                "row": tc_entry["row"],
                "source_sheet": "Test Case Remarks"
            })
    
    if not final_results:
        raise HTTPException(status_code=400, detail="No valid data found in either Test Scenario Remarks or Test Case Remarks sheets.")
    
    return final_results
