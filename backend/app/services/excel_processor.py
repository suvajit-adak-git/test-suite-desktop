import re
from typing import Dict, Optional, List
from pathlib import Path
from openpyxl import load_workbook

class ExcelHyperlinkProcessor:
    """Process and update hyperlinks in Excel files with build numbers."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.hyperlinks_data = []
        
    def load_workbook(self):
        """Load the Excel workbook."""
        try:
            self.workbook = load_workbook(self.file_path)
        except Exception as e:
            raise Exception(f"Failed to load workbook: {str(e)}")
    
    def extract_hyperlinks(self) -> Dict:
        """
        Extract hyperlinks from 'Test Case Remarks' and 'Test Scenario Remarks' sheets.
        Returns JSON with file names, addresses, and build numbers.
        """
        if not self.workbook:
            self.load_workbook()
        
        required_sheets = ['Test Case Remarks', 'Test Scenario Remarks']
        self.hyperlinks_data = []
        errors = []
        
        for sheet_name in required_sheets:
            if sheet_name not in self.workbook.sheetnames:
                errors.append(f"Sheet '{sheet_name}' not found in the workbook")
                continue
            
            sheet = self.workbook[sheet_name]
            
            # Iterate through all cells to find hyperlinks
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        hyperlink_address = cell.hyperlink.target
                        
                        # Extract build number
                        build_match = re.search(r'_0(\d+)', hyperlink_address)
                        
                        if build_match:
                            build_number = build_match.group(1)  # Gets digits after _0
                            
                            # Extract file name from URL
                            file_name = hyperlink_address.split('/')[-1] if '/' in hyperlink_address else hyperlink_address
                            
                            hyperlink_info = {
                                'sheet_name': sheet_name,
                                'cell_reference': cell.coordinate,
                                'file_name': file_name,
                                'file_address': hyperlink_address,
                                'build': build_number,
                                'full_build_pattern': f"_0{build_number}"
                            }
                            
                            self.hyperlinks_data.append(hyperlink_info)
                        else:
                            # Check if there's a build pattern without leading zero
                            invalid_build = re.search(r'_(\d+)(?!/)', hyperlink_address)
                            if invalid_build and not invalid_build.group(0).startswith('_0'):
                                errors.append({
                                    'sheet': sheet_name,
                                    'cell': cell.coordinate,
                                    'address': hyperlink_address,
                                    'error': f"Invalid build pattern found: {invalid_build.group(0)}. Build must have leading zero (e.g., _0212)"
                                })
        
        result = {
            'status': 'success' if not errors else 'partial_success',
            'total_hyperlinks': len(self.hyperlinks_data),
            'hyperlinks': self.hyperlinks_data,
            'errors': errors if errors else None
        }
        
        return result
    
    def update_build_numbers(self, new_build: str, output_file_path: Optional[str] = None) -> Dict:
        """
        Update all hyperlinks with new build number.
        
        Args:
            new_build: New build number (e.g., '213')
            output_file_path: Path for the updated file. If None, generates automatically.
        
        Returns:
            JSON with update status and details.
        """
        if not self.workbook:
            self.load_workbook()
        
        # Ensure build number doesn't have leading underscore or zero
        new_build = new_build.lstrip('_0')
        new_build_pattern = f"_0{new_build}"
        
        required_sheets = ['Test Case Remarks', 'Test Scenario Remarks']
        updated_count = 0
        update_details = []
        
        for sheet_name in required_sheets:
            if sheet_name not in self.workbook.sheetnames:
                continue
            
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        old_address = cell.hyperlink.target
                        
                        # Find and replace build pattern
                        build_match = re.search(r'_0\d+', old_address)
                        
                        if build_match:
                            old_build_pattern = build_match.group(0)
                            new_address = old_address.replace(old_build_pattern, new_build_pattern)
                            
                            # Update the hyperlink
                            cell.hyperlink.target = new_address
                            
                            updated_count += 1
                            update_details.append({
                                'sheet': sheet_name,
                                'cell': cell.coordinate,
                                'old_address': old_address,
                                'new_address': new_address,
                                'old_build': old_build_pattern.lstrip('_0'),
                                'new_build': new_build
                            })
        
        # Save to new file
        if not output_file_path:
            file_path = Path(self.file_path)
            output_file_path = file_path.parent / f"{file_path.stem}_build_{new_build}{file_path.suffix}"
        
        try:
            self.workbook.save(output_file_path)
            status = 'success'
            message = f"Successfully updated {updated_count} hyperlinks"
        except Exception as e:
            status = 'error'
            message = f"Failed to save updated file: {str(e)}"
            output_file_path = None
        
        result = {
            'status': status,
            'message': message,
            'new_build': new_build,
            'updated_count': updated_count,
            'output_file': str(output_file_path) if output_file_path else None,
            'updates': update_details
        }
        
        return result
    
    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()
