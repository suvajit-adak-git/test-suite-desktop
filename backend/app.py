# app.py
import shutil
import tempfile
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
import os
import re
import json
import difflib

from fastapi import FastAPI, UploadFile, File, HTTPException, Body
from pydantic import BaseModel

import pandas as pd
import openpyxl

app = FastAPI(title="Inspector: SVN CSV + Review Checklist (with advanced compare)")

UPLOAD_DIR = Path(tempfile.gettempdir()) / "uploads_backend"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

def save_upload_file(upload_file: UploadFile, dest: Path) -> None:
    with dest.open("wb") as buffer:
        shutil.copyfileobj(upload_file.file, buffer)

# ---------- helpers: csv/excel extractors (unchanged) ----------
def extract_from_excel_file(path: Path, max_preview_rows: int = 100) -> Dict[str, Any]:
    try:
        df = pd.read_excel(path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read excel file: {e}")

    headers = list(df.columns)
    nrows = int(len(df))
    preview_df = df.head(max_preview_rows).fillna("")
    preview = preview_df.to_dict(orient="records")
    sample_values = []
    if headers:
        first_col = headers[0]
        sample_values = [str(x).strip() for x in df[first_col].dropna().astype(str).unique().tolist()]

    return {"filename": path.name, "headers": headers, "nrows": nrows, "preview": preview, "sample_values": sample_values}

def extract_from_csv_file(path: Path, max_preview_rows: int = 100) -> Dict[str, Any]:
    df = None
    try:
        df = pd.read_csv(path, engine="python", sep=None)
    except Exception:
        for sep in [",", ";", "\t", "|"]:
            try:
                df = pd.read_csv(path, sep=sep)
                break
            except Exception:
                df = None
    if df is None:
        raise HTTPException(status_code=400, detail="Failed to parse CSV file (invalid format or delimiter).")

    df = df.dropna(how="all")
    df = df[~(df.astype(str).apply(lambda x: "".join(x).strip(), axis=1) == "")]

    headers = list(df.columns)
    nrows = int(len(df))
    preview_df = df.head(max_preview_rows).fillna("")
    preview = preview_df.to_dict(orient="records")
    sample_values = []
    if headers:
        first_col = headers[0]
        sample_values = [str(x).strip() for x in df[first_col].dropna().astype(str).unique().tolist()]

    return {"filename": path.name, "headers": headers, "nrows": nrows, "preview": preview, "sample_values": sample_values}

# ---------- review checklist extractor ----------
def _cell_fill_rgb(cell) -> Optional[str]:
    try:
        f = cell.fill
        if not f:
            return None
        color = f.start_color
        if color is None:
            return None
        rgb = getattr(color, "rgb", None)
        if rgb:
            return str(rgb).upper()
        return None
    except Exception:
        return None

def extract_hyperlinks_with_versions_from_path(file_path: str, sheet_name: str = "Test Scenario Remarks") -> List[Dict[str, Any]]:
    p = Path(file_path)
    if not p.exists() or not p.is_file():
        raise HTTPException(status_code=404, detail=f"File not found: {file_path}")

    try:
        workbook = openpyxl.load_workbook(p, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to open workbook: {e}")

    if sheet_name not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found. Available: {', '.join(workbook.sheetnames)}")

    sheet = workbook[sheet_name]

    filename_col = None
    version_closed_col = None
    header_row = None

    for row_idx in range(1, min(21, sheet.max_row + 1)):
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row_idx, col_idx).value
            if cell_value:
                cell_value_str = str(cell_value).strip().lower()
                if "filename" in cell_value_str:
                    filename_col = col_idx
                    header_row = row_idx
                if "version on which" in cell_value_str and "closed" in cell_value_str:
                    version_closed_col = col_idx
        if filename_col and version_closed_col:
            break

    if not filename_col or not version_closed_col:
        raise HTTPException(status_code=400, detail="Required columns not found ('Filename' and 'Version on which review closed').")

    results: List[Dict[str, Any]] = []

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        filename_cell = sheet.cell(row_idx, filename_col)
        version_cell = sheet.cell(row_idx, version_closed_col)

        color = _cell_fill_rgb(filename_cell)
        if color and len(color) >= 6:
            hexstr = color[-8:] if len(color) >= 8 else color
            try:
                if len(hexstr) == 8:
                    r = int(hexstr[2:4], 16); g = int(hexstr[4:6], 16); b = int(hexstr[6:8], 16)
                elif len(hexstr) == 6:
                    r = int(hexstr[0:2], 16); g = int(hexstr[2:4], 16); b = int(hexstr[4:6], 16)
                else:
                    r = g = b = 0
            except Exception:
                r = g = b = 0

            if g > r and g > b and g > 100:
                break

        filename = filename_cell.value
        version_closed = version_cell.value

        if not filename:
            continue

        hyperlink_url = None
        if getattr(filename_cell, "hyperlink", None):
            try:
                hyperlink_url = filename_cell.hyperlink.target
            except Exception:
                hyperlink_url = None

        filename_str = str(filename).strip() if filename is not None else None
        version_closed_str = str(version_closed).strip() if version_closed is not None else None

        results.append({
            "filename": filename_str,
            "hyperlink": hyperlink_url,
            "version_closed": version_closed_str,
            "row": row_idx
        })

    return results

# ---------- Utility: normalization and comparison helpers ----------

def _strip_extension(name: str) -> str:
    # Path.stem removes extension, but for names that include paths we use basename then stem
    try:
        return Path(name).stem
    except Exception:
        # fallback: remove last dot extension
        if "." in name:
            return ".".join(name.split(".")[:-1])
        return name

def _normalize_filename_for_match(name: Optional[str]) -> str:
    if not name:
        return ""
    s = str(name).strip().lower()
    # remove file extension
    s = _strip_extension(s)
    # replace non-alphanumeric with space, collapse whitespace
    s = re.sub(r"[^0-9a-z]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _extract_int_from_version(s: Optional[str]) -> Optional[int]:
    """
    Try to extract integer version from strings like 'v20157', ' 20157 ', '20157.0' etc.
    Returns int or None if not parseable.
    """
    if s is None:
        return None
    s = str(s).strip()
    # find first group of digits
    m = re.search(r"(\d+)", s)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None

def _normalize_version_string(s: Optional[str]) -> str:
    if s is None:
        return ""
    return str(s).strip()

def _fuzzy_best_match(key: str, candidates: List[str]) -> Tuple[Optional[str], float]:
    """
    Return (best_candidate, score) using difflib.SequenceMatcher ratio.
    Score is between 0..1. If no candidates, returns (None, 0.0)
    """
    if not candidates:
        return None, 0.0
    best = None
    best_score = 0.0
    for c in candidates:
        score = difflib.SequenceMatcher(None, key, c).ratio()
        if score > best_score:
            best_score = score
            best = c
    return best, best_score

# ---------- compare endpoint with enhanced logic ----------

@app.post("/api/compare-both")
async def compare_both(payload: Dict[str, Any] = Body(...)):
    """
    Compare svn blob and checklist blob (or use server-local paths).
    Enhancements:
      - filename normalization + strip extension
      - version coercion to integer when possible
      - fuzzy matching (difflib) with threshold (default 0.85)
    Payload options:
      - { "svn": <svn_blob>, "checklist": <checklist_blob>, "fuzzy_threshold": 0.85 }
      - or provide { "svn_path": "/abs/path/to/svn_report.csv", "checklist_path": "/abs/path/to/checklist.xlsx", "sheet_name": "...", "fuzzy_threshold": 0.85 }
    """
    fuzzy_threshold = float(payload.get("fuzzy_threshold", 0.85))

    svn_blob = payload.get("svn")
    checklist_blob = payload.get("checklist")

    # If server-local paths provided, process them first
    if not svn_blob and payload.get("svn_path"):
        svn_path = payload["svn_path"]
        # accept file:// URIs by stripping
        if svn_path.startswith("file://"):
            svn_path = svn_path[len("file://"):]
        if not os.path.isabs(svn_path):
            raise HTTPException(status_code=400, detail="svn_path must be absolute")
        svn_path_obj = Path(svn_path)
        if not svn_path_obj.exists():
            raise HTTPException(status_code=404, detail=f"svn_path not found: {svn_path}")
        if svn_path_obj.suffix.lower() == ".csv":
            svn_blob = extract_from_csv_file(svn_path_obj)
        else:
            svn_blob = extract_from_excel_file(svn_path_obj)

    if not checklist_blob and payload.get("checklist_path"):
        checklist_path = payload["checklist_path"]
        if checklist_path.startswith("file://"):
            checklist_path = checklist_path[len("file://"):]
        if not os.path.isabs(checklist_path):
            raise HTTPException(status_code=400, detail="checklist_path must be absolute")
        if not Path(checklist_path).exists():
            raise HTTPException(status_code=404, detail=f"checklist_path not found: {checklist_path}")
        sheet_name = payload.get("sheet_name", "Test Scenario Remarks")
        checklist_list = extract_hyperlinks_with_versions_from_path(checklist_path, sheet_name=sheet_name)
        checklist_blob = {"filename": Path(checklist_path).name, "data": checklist_list, "count": len(checklist_list)}

    if not svn_blob or not checklist_blob:
        raise HTTPException(status_code=400, detail="Provide either 'svn' and 'checklist' blobs, or 'svn_path' and 'checklist_path'.")

    # Extract svn rows list
    if isinstance(svn_blob, dict) and "preview" in svn_blob:
        svn_rows = svn_blob["preview"]
    elif isinstance(svn_blob, dict) and "data" in svn_blob:
        svn_rows = svn_blob["data"]
    elif isinstance(svn_blob, list):
        svn_rows = svn_blob
    else:
        raise HTTPException(status_code=400, detail="Unrecognized svn blob format")

    # Extract checklist rows list
    if isinstance(checklist_blob, dict) and "data" in checklist_blob:
        checklist_rows = checklist_blob["data"]
    elif isinstance(checklist_blob, list):
        checklist_rows = checklist_blob
    else:
        raise HTTPException(status_code=400, detail="Unrecognized checklist blob format")

    # Build canonical maps keyed by normalized filename (no extension)
    svn_map: Dict[str, Dict[str, Any]] = {}
    for r in svn_rows:
        # support both "File" or lowercase/other keys
        filename = r.get("File") if "File" in r else r.get("file") or r.get("Filename") or r.get("filename")
        if not filename:
            continue
        norm = _normalize_filename_for_match(filename)
        svn_rev_raw = r.get("Last Changed Revision") or r.get("Last Changed Revision".lower()) or r.get("WC Revision") or r.get("revision") or r.get("Revision")
        svn_auth = r.get("Last Changed Author") or r.get("Last Changed Author".lower()) or r.get("last changed author") or r.get("Last Changed Author")
        svn_date = r.get("Last Changed Date") or r.get("Last Changed Date".lower()) or r.get("last changed date")
        svn_map[norm] = {
            "raw": r,
            "norm_name": norm,
            "filename_original": filename,
            "last_changed_revision_raw": _normalize_version_string(svn_rev_raw),
            "last_changed_revision_int": _extract_int_from_version(svn_rev_raw),
            "last_changed_author": svn_auth,
            "last_changed_date": svn_date
        }

    checklist_map: Dict[str, Dict[str, Any]] = {}
    checklist_norm_list: List[str] = []
    for r in checklist_rows:
        filename = r.get("filename") or r.get("Filename") or r.get("File")
        if not filename:
            continue
        norm = _normalize_filename_for_match(filename)
        checklist_norm_list.append(norm)
        checklist_map[norm] = {
            "raw": r,
            "norm_name": norm,
            "filename_original": filename,
            "version_closed_raw": _normalize_version_string(r.get("version_closed") or r.get("Version") or r.get("version")),
            "version_closed_int": _extract_int_from_version(r.get("version_closed") or r.get("Version") or r.get("version"))
        }

    # Matching process
    matches = []
    mismatches = []
    only_in_svn = []
    only_in_checklist = []

    used_checklist_keys = set()

    # First pass: exact normalized matches
    for s_key, s in svn_map.items():
        if s_key in checklist_map:
            c = checklist_map[s_key]
            used_checklist_keys.add(s_key)

            # compare versions (prefer integer comparison when both available)
            s_ver_int = s["last_changed_revision_int"]
            c_ver_int = c["version_closed_int"]
            s_ver_raw = s["last_changed_revision_raw"]
            c_ver_raw = c["version_closed_raw"]

            version_equal = False
            if s_ver_int is not None and c_ver_int is not None:
                version_equal = (s_ver_int == c_ver_int)
            else:
                # fallback to trimmed string comparison
                version_equal = (s_ver_raw != "" and c_ver_raw != "" and s_ver_raw == c_ver_raw)

            entry = {
                "filename": s["filename_original"],
                "normalized_filename": s_key,
                "svn_revision_raw": s_ver_raw,
                "svn_revision_int": s_ver_int,
                "checklist_version_raw": c_ver_raw,
                "checklist_version_int": c_ver_int,
                "last_changed_author": s["last_changed_author"],
                "last_changed_date": s["last_changed_date"],
                "match_type": "exact",
                "score": 1.0
            }
            if version_equal:
                matches.append(entry)
            else:
                mismatches.append(entry)
        else:
            # not exact match; handle later with fuzzy
            pass

    # Second pass: for svn keys not matched, try fuzzy matching
    svn_unmatched = [k for k in svn_map.keys() if k not in {m["normalized_filename"] for m in matches + mismatches}]
    checklist_candidates = [k for k in checklist_map.keys() if k not in used_checklist_keys]

    for s_key in svn_unmatched:
        s = svn_map[s_key]
        best_candidate, score = _fuzzy_best_match(s_key, checklist_candidates)
        if best_candidate and score >= fuzzy_threshold:
            # accept fuzzy match
            c = checklist_map[best_candidate]
            used_checklist_keys.add(best_candidate)
            checklist_candidates.remove(best_candidate)

            s_ver_int = s["last_changed_revision_int"]
            c_ver_int = c["version_closed_int"]
            s_ver_raw = s["last_changed_revision_raw"]
            c_ver_raw = c["version_closed_raw"]

            version_equal = False
            if s_ver_int is not None and c_ver_int is not None:
                version_equal = (s_ver_int == c_ver_int)
            else:
                version_equal = (s_ver_raw != "" and c_ver_raw != "" and s_ver_raw == c_ver_raw)

            entry = {
                "filename": s["filename_original"],
                "normalized_filename": s_key,
                "matched_checklist_filename": c["filename_original"],
                "matched_checklist_normalized": best_candidate,
                "svn_revision_raw": s_ver_raw,
                "svn_revision_int": s_ver_int,
                "checklist_version_raw": c_ver_raw,
                "checklist_version_int": c_ver_int,
                "last_changed_author": s["last_changed_author"],
                "last_changed_date": s["last_changed_date"],
                "match_type": "fuzzy",
                "score": score
            }
            if version_equal:
                matches.append(entry)
            else:
                mismatches.append(entry)
        else:
            # no match found
            only_in_svn.append({
                "filename": s["filename_original"],
                "normalized_filename": s_key,
                "last_changed_revision_raw": s["last_changed_revision_raw"],
                "last_changed_revision_int": s["last_changed_revision_int"],
                "last_changed_author": s["last_changed_author"],
                "last_changed_date": s["last_changed_date"]
            })

    # Any checklist keys left unused are only_in_checklist
    for c_key, c in checklist_map.items():
        if c_key not in used_checklist_keys:
            only_in_checklist.append({
                "filename": c["filename_original"],
                "normalized_filename": c_key,
                "version_closed_raw": c["version_closed_raw"],
                "version_closed_int": c["version_closed_int"],
                "raw": c["raw"]
            })

    return {
        "status": "ok",
        "summary": {
            "matches": len(matches),
            "mismatches": len(mismatches),
            "only_in_svn": len(only_in_svn),
            "only_in_checklist": len(only_in_checklist)
        },
        "matches": matches,
        "mismatches": mismatches,
        "only_in_svn": only_in_svn,
        "only_in_checklist": only_in_checklist
    }

# ---------- Existing upload endpoints (unchanged) ----------

@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    fname = file.filename or ""
    fname_lower = fname.lower()
    if not fname_lower.endswith((".xls", ".xlsx", ".csv")):
        raise HTTPException(status_code=400, detail="Please upload an .xls, .xlsx or .csv file")

    dest = UPLOAD_DIR / fname
    counter = 0
    base = dest.stem
    ext = dest.suffix
    while dest.exists():
        counter += 1
        dest = UPLOAD_DIR / f"{base}_{counter}{ext}"

    try:
        save_upload_file(file, dest)
        if dest.suffix.lower() == ".csv":
            data = extract_from_csv_file(dest)
        else:
            data = extract_from_excel_file(dest)
    finally:
        try:
            dest.unlink()
        except Exception:
            pass

    return {"status": "ok", "data": data}

@app.post("/api/upload-review-checklist")
async def upload_review_checklist(file: UploadFile = File(...), sheet_name: Optional[str] = "Test Scenario Remarks"):
    fname = file.filename or ""
    if not fname.lower().endswith((".xls", ".xlsx")):
        raise HTTPException(status_code=400, detail="Please upload an .xls or .xlsx file for review checklist")

    dest = UPLOAD_DIR / fname
    counter = 0
    base = dest.stem
    ext = dest.suffix
    while dest.exists():
        counter += 1
        dest = UPLOAD_DIR / f"{base}_{counter}{ext}"

    try:
        save_upload_file(file, dest)
        results = extract_hyperlinks_with_versions_from_path(str(dest), sheet_name=sheet_name)
    finally:
        try:
            dest.unlink()
        except Exception:
            pass

    return {"status": "ok", "data": results, "count": len(results)}

@app.post("/api/upload-both")
async def upload_both(svn_file: UploadFile = File(...), checklist_file: UploadFile = File(...), sheet_name: Optional[str] = "Test Scenario Remarks"):
    svn_name = svn_file.filename or "svn_input"
    svn_dest = UPLOAD_DIR / svn_name
    i = 0
    while svn_dest.exists():
        i += 1
        svn_dest = UPLOAD_DIR / f"{Path(svn_name).stem}_{i}{svn_dest.suffix}"

    check_name = checklist_file.filename or "checklist_input.xlsx"
    check_dest = UPLOAD_DIR / check_name
    j = 0
    while check_dest.exists():
        j += 1
        check_dest = UPLOAD_DIR / f"{Path(check_name).stem}_{j}{check_dest.suffix}"

    try:
        save_upload_file(svn_file, svn_dest)
        save_upload_file(checklist_file, check_dest)

        if svn_dest.suffix.lower() == ".csv":
            svn_data = extract_from_csv_file(svn_dest)
        else:
            svn_data = extract_from_excel_file(svn_dest)

        checklist_data = extract_hyperlinks_with_versions_from_path(str(check_dest), sheet_name=sheet_name)
    finally:
        try:
            svn_dest.unlink()
        except Exception:
            pass
        try:
            check_dest.unlink()
        except Exception:
            pass

    return {"status": "ok", "svn": svn_data, "checklist": {"filename": check_name, "data": checklist_data, "count": len(checklist_data)}}

class LocalPathsRequest(BaseModel):
    svn_path: str
    checklist_path: str
    sheet_name: Optional[str] = "Test Scenario Remarks"

@app.post("/api/process-local-paths")
async def process_local_paths(req: LocalPathsRequest):
    if not req.svn_path or not os.path.isabs(req.svn_path):
        raise HTTPException(status_code=400, detail="Provide an absolute svn_path on the server")
    if not req.checklist_path or not os.path.isabs(req.checklist_path):
        raise HTTPException(status_code=400, detail="Provide an absolute checklist_path on the server")

    if not Path(req.svn_path).exists():
        raise HTTPException(status_code=404, detail=f"svn_path not found: {req.svn_path}")
    if not Path(req.checklist_path).exists():
        raise HTTPException(status_code=404, detail=f"checklist_path not found: {req.checklist_path}")

    svn_path_obj = Path(req.svn_path)
    if svn_path_obj.suffix.lower() == ".csv":
        svn_data = extract_from_csv_file(svn_path_obj)
    else:
        svn_data = extract_from_excel_file(svn_path_obj)

    checklist_data = extract_hyperlinks_with_versions_from_path(req.checklist_path, sheet_name=req.sheet_name)

    return {"status": "ok", "svn": svn_data, "checklist": {"filename": Path(req.checklist_path).name, "data": checklist_data, "count": len(checklist_data)}}

@app.get("/health")
async def health():
    return {"status": "ok"}
