import pytest
import os
from openpyxl import Workbook
from app.services.excel_processor import ExcelHyperlinkProcessor

@pytest.fixture
def sample_excel_file(tmp_path):
    wb = Workbook()
    
    # Create Test Case Remarks sheet
    ws1 = wb.active
    ws1.title = "Test Case Remarks"
    ws1['A1'] = "Link"
    ws1['A1'].hyperlink = "http://example.com/file_0123.txt"
    ws1['A2'] = "No Link"
    
    # Create Test Scenario Remarks sheet
    ws2 = wb.create_sheet("Test Scenario Remarks")
    ws2['B2'] = "Another Link"
    ws2['B2'].hyperlink = "http://example.com/other_0456.doc"
    
    file_path = tmp_path / "test_hyperlinks.xlsx"
    wb.save(file_path)
    return str(file_path)

def test_extract_hyperlinks(sample_excel_file):
    processor = ExcelHyperlinkProcessor(sample_excel_file)
    result = processor.extract_hyperlinks()
    processor.close()
    
    assert result['status'] == 'success'
    assert result['total_hyperlinks'] == 2
    
    hyperlinks = result['hyperlinks']
    assert len(hyperlinks) == 2
    
    # Check first hyperlink
    link1 = next(l for l in hyperlinks if l['build'] == '123')
    assert link1['sheet_name'] == 'Test Case Remarks'
    assert link1['file_name'] == 'file_0123.txt'
    
    # Check second hyperlink
    link2 = next(l for l in hyperlinks if l['build'] == '456')
    assert link2['sheet_name'] == 'Test Scenario Remarks'
    assert link2['file_name'] == 'other_0456.doc'

def test_update_build_numbers(sample_excel_file):
    processor = ExcelHyperlinkProcessor(sample_excel_file)
    new_build = "789"
    result = processor.update_build_numbers(new_build)
    processor.close()
    
    assert result['status'] == 'success'
    assert result['updated_count'] == 2
    assert result['new_build'] == "789"
    
    # Verify the updated file
    output_file = result['output_file']
    assert os.path.exists(output_file)
    
    # Load updated file and check hyperlinks
    processor_updated = ExcelHyperlinkProcessor(output_file)
    result_updated = processor_updated.extract_hyperlinks()
    processor_updated.close()
    
    hyperlinks = result_updated['hyperlinks']
    for link in hyperlinks:
        assert link['build'] == "789"
        assert "_0789" in link['file_address']

def test_invalid_sheet(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Wrong Sheet"
    file_path = tmp_path / "wrong_sheet.xlsx"
    wb.save(file_path)
    
    processor = ExcelHyperlinkProcessor(str(file_path))
    result = processor.extract_hyperlinks()
    processor.close()
    
    assert result['status'] == 'partial_success' # Or success with errors, depending on implementation details. Code says partial_success if errors.
    assert result['errors'] is not None
    assert len(result['errors']) == 2 # Both required sheets missing
